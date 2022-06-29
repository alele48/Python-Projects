from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import os
import time
import re

"""
The following program utilizes the openpyxl module to
automate excel staff payroll.
"""

def excel_file(ws):

    while True:
        try:
            # asks for number of employee
            prompt = int(input("How many employee are you going to input: "))
            if prompt > 25:
                print("Too large")
            elif prompt < 1:
                print("Too small")
            else:
                break
        except:
            print("Oops! Please enter a number")
            time.sleep(1.5)

    i = 0
    for cycle in range(prompt):
        name = input("Employee Name: ")
        dep = input("Department: ")
        hour = float(input("Hours: "))
        hp = float(input("Hourly Pay: "))
        tax = float("%1".strip("%"))
        pay_before_tax = hour * hp
        reduction = pay_before_tax * ( tax * 0.01)
        total = pay_before_tax - reduction
        ws.append([name, dep, hour, hp, tax, total])


def new_workbook():
    wb = Workbook()
    ws = wb.active # gets active worksheet from workbook
    worksheet = input("Name worksheet: ").upper()
    ws.title = worksheet
    ws.append(["Name", "Department", "Hour", "Hourly Pay", "Tax", "Total"])
    # returns a-g at column 1
    for row in range(1,2):
        for col in range(1,8):
            char = get_column_letter(col)
            # first row turns bold
            ws[char + str(row)].font = Font(bold=True)
    while True:
        workbook = input("Name Your file: ").upper()+".xlsx"
        path = os.getcwd()
        files_folders = os.listdir(path)
        if workbook in files_folders:
            print("File already exists")
            time.sleep(1)
        else:
            excel_file(ws)
            wb.save(workbook)
            break

def load_work():
    # gets current working directory
    current_dir = os.getcwd()
    res = []
    for path in os.listdir(current_dir):
        # check for files
        if os.path.isfile(os.path.join(current_dir, path)):
            res.append(path)

    # checks if list contains excel file
    # and prints them out with xlsx extension

    for r in res:
        if r.endswith("xlsx"):
            print(r.strip(".xlsx"))

    while True:
        chosen_wb = input("Choose from listed files: ").upper()+".xlsx"
        if chosen_wb not in res:
            print("File does not exist")
            time.sleep(1)
        else:
            while True:
                choice = input("Press 1 to create new sheet, 2 to add to existing sheet, b go back to menu, q to quit: ").upper()
                if choice == "1":
                    wb = load_workbook(chosen_wb) #opens existing workbook
                    worksheet = input("Name worksheet: ").upper()
                    ws = wb.create_sheet(worksheet) # creates new sheet
                    ws.append(["Name", "Department", "Hour", "Hourly Pay", "Tax", "Total"])
                    # returns a-g at column 1
                    for row in range(1,2):
                        for col in range(1,8):
                            char = get_column_letter(col)
                            # first row turns bold
                            ws[char + str(row)].font = Font(bold=True)
                    excel_file(ws)
                    wb.save(chosen_wb)
                    break
                elif choice == "2":
                    wb = load_workbook(chosen_wb)
                    ws_list = []
                    for sheet in wb.worksheets:
                        
                        sheet = str(sheet)
                        sheet = re.sub("[<>Worksheet]",'', sheet)
                        sheet = sheet.replace('"','')
                        print(sheet)
                        ws_list.append(sheet)
                    while True:
                        try:
                            chosen_ws = input("Choose sheet: ").upper()
                            wb.active = wb[chosen_ws]
                            ws = wb.active
                            excel_file(ws)
                            wb.save(chosen_wb)  
                            break
                        except:
                            print("Oops! Wrong file. Try again.")
                elif choice == 'Q':
                    quit()
                elif choice == 'B':
                    main()
                    
                else:
                    print("Press 1 or 2")
                    time.sleep(1)
            break

# main menu loops over the program
def main():
    while True:
        select = input("Press 1 to create new file, 2 to load file, q to quit: ").upper()
        if select == "1":
            new_workbook()
        elif select == "2":
            load_work()
        elif select == "Q":
            exit()
        else:
            print("Press 1, 2, or q: ")
            time.sleep(1)
main()
